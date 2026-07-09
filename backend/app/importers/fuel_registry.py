"""Importer for fuel-card transaction exports ("Топливо").

Built 2026-06-19 evening against a real E100 export (fuel_transfer_*.xlsx,
6 data rows). Column layout verified against that fixture:

    ID | Дата | АЗС | Карта № | Марка ТС | Гос. Номер ТС | Объем | Сумма | ID транзакции

"Дата" uses a comma between date and time ("03.06.2026, 10:42:38"), unlike
the trip registry's format - hence the explicit `fmt` passed to
common._parse_dt() below. Other fuel-card providers may use a different
layout; that's deferred until a second example shows up (same pattern as
trip_registry.py's OZON-only EXPECTED_COLUMNS).

Rows are upserted by the file's own "ID" column (`external_id`) so
re-importing an overlapping export doesn't duplicate transactions. Plate
matching reuses the same normalize_plate()/find_or_create_truck() as the
trip importer (see importers/common.py) so a truck created from a trip
import and a truck created from a fuel import converge on one row.
"""

from io import BytesIO

import openpyxl
from sqlmodel import Session, select

from .. import models
from .common import _cell_str, _parse_dt, find_or_create_truck, normalize_plate

FUEL_DATE_FORMAT = "%d.%m.%Y, %H:%M:%S"

FUEL_COLUMNS = {
    "external_id": "ID",
    "date": "Дата",
    "station": "АЗС",
    "card_number": "Карта №",
    "truck_brand": "Марка ТС",
    "plate": "Гос. Номер ТС",
    "volume": "Объем",
    "amount": "Сумма",
    "external_transaction_id": "ID транзакции",
}


def import_fuel_records(file_bytes: bytes, session: Session) -> dict:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_index = {str(name).strip(): idx for idx, name in enumerate(header_row) if name}

    def col(row, key):
        idx = header_index.get(FUEL_COLUMNS[key])
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    existing = {
        r.external_id: r
        for r in session.exec(select(models.FuelRecord)).all()
        if r.external_id
    }
    truck_cache: dict = {}
    new_truck_labels = []

    records_created = 0
    records_updated = 0
    skipped_bad_rows = 0

    for row in rows:
        if row is None or all(c is None for c in row):
            continue

        external_id = _cell_str(col(row, "external_id"))
        if not external_id:
            skipped_bad_rows += 1
            continue

        try:
            record_date = _parse_dt(col(row, "date"), fmt=FUEL_DATE_FORMAT)
        except (ValueError, TypeError):
            record_date = None
        if record_date is None:
            skipped_bad_rows += 1
            continue

        plate_raw = _cell_str(col(row, "plate"))

        truck_id = None
        if plate_raw:
            norm_plate = normalize_plate(plate_raw)
            if norm_plate not in truck_cache:
                truck, was_created = find_or_create_truck(session, plate_raw)
                truck_cache[norm_plate] = truck
                if was_created:
                    new_truck_labels.append(truck.label)
            truck_id = truck_cache[norm_plate].id

        fields = dict(
            external_id=external_id,
            date=record_date,
            station=_cell_str(col(row, "station")),
            card_number=_cell_str(col(row, "card_number")),
            truck_brand_raw=_cell_str(col(row, "truck_brand")),
            plate_raw=plate_raw,
            truck_id=truck_id,
            volume=col(row, "volume") or 0,
            amount=col(row, "amount") or 0,
            external_transaction_id=_cell_str(col(row, "external_transaction_id")),
        )

        record = existing.get(external_id)
        if record:
            for k, v in fields.items():
                setattr(record, k, v)
            records_updated += 1
        else:
            record = models.FuelRecord(**fields)
            existing[external_id] = record
            records_created += 1
        session.add(record)

    session.commit()

    return {
        "total_rows": len(rows),
        "records_created": records_created,
        "records_updated": records_updated,
        "skipped_bad_rows": skipped_bad_rows,
        "new_trucks": new_truck_labels,
    }

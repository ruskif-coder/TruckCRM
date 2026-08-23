"""Importer for the per-trip registry export ("реестр поездок").

Two import paths live in this module:

  import_trips()          - LIVE path (redesigned 2026-06-19). Parses every
                             row into a raw, ungrouped models.Trip - one row
                             per source line, every status included (even
                             "Отменено"). This is what the API/frontend use
                             today: a flat, sortable trip list.

  import_trip_registry()  - DORMANT path, kept as-is, NOT called from the
                             live import flow anymore. Aggregates rows into
                             weekly models.TripBatch records (rule confirmed
                             2026-06-19 against the real OZON weekly report,
                             TCargo 2025.xlsx, matched all 4 rows to the
                             ruble - see memory: transport-crm-rebuild-scope).
                             Reserved for the future dashboard feature; the
                             user explicitly asked to preserve this logic
                             rather than delete it when the raw-list import
                             was introduced.

      1. Exclude rows with Статус заявки == "Отменено" (cancelled). Both
         "Получен ответ" and "Завершено" are kept.
      2. Group by ISO calendar week (Mon-Sun) of "Дата отгрузки" + driver.
      3. period_start/period_end = min/max trip date within the (week,
         driver) group (not the fixed week boundary) - a mid-week driver
         handover produces two separate batches instead of one.
      4. trips_count = row count in the group; gross_revenue = sum of
         "Сумма транзакций" in the group.

Column layout below is OZON's specific export. Other carriers may use a
different layout - that's deferred until a second example file shows up
(see memory: transport-crm-rebuild-scope).

Driver-name matching: the file gives "Отчество Имя Фамилия" while the app
stores "Фамилия Имя". Matching is done on the surname token; if nothing
matches, a new Driver (and, if needed, a new Truck by plate) is created
automatically (decided 2026-06-20) so the import never blocks on missing
reference data. Same matching/auto-create is reused by import_trips().
"""

from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlmodel import Session, select

from .. import models
from .common import (
    _cell_str,
    _parse_dt,
    _surname_token,
    find_or_create_driver,
    find_or_create_truck,
    normalize_plate,
)

STATUS_EXCLUDE = {"Отменено"}

EXPECTED_COLUMNS = {
    "status": "Статус заявки",
    "plate": "гос. Номер ТС",
    "driver": "ФИО водителя",
    "dep_date": "Дата отгрузки (Часовой пояс точки)",
    "amount": "Сумма транзакций",
    "trip_type": "Тип тарификации",
}

# Full column map for the raw/ungrouped import (import_trips). Verified
# 2026-06-19 against the real реестр поездок.xlsx (59 rows): the
# confirmation-date header contains a literal backslash, not a slash.
TRIP_COLUMNS = {
    "request_number": "№ заявки",
    "external_request_number": "Внешний № заявки",
    "tariff_type": "Тип тарификации",
    "plate": "гос. Номер ТС",
    "driver": "ФИО водителя",
    "driver_phone": "Телефон водителя",
    "confirmed_at": "Дата\\время подтверждения заявки (МСК)",
    "status": "Статус заявки",
    "dep_at": "Дата отгрузки (Часовой пояс точки)",
    "end_at": "Дата окончания рейса",
    "amount": "Сумма транзакций",
    # Added 2026-06-26: the user added this column to their copy of the
    # import template by hand (see models.TripBase.fines). col() below
    # returns None when a file doesn't have this header, which import_trips
    # already treats as "no value" (falls back to 0) - so older files
    # without this column still import fine.
    "fines": "Штраф",
}


def build_import_template() -> bytes:
    """Blank .xlsx with the exact header row import_trips() expects, plus one
    example row showing the expected value shapes (dates as real Excel dates,
    not text) - added 2026-06-26 so the user has something to hand a carrier
    instead of reverse-engineering TRIP_COLUMNS from a real OZON export.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Реестр поездок"

    headers = list(TRIP_COLUMNS.values())
    ws.append(headers)
    ws.append(
        [
            "12345",  # № заявки
            "EXT-12345",  # Внешний № заявки
            "Городские перевозки",  # Тип тарификации
            "А123ВС77",  # гос. Номер ТС
            "Иванович Иван Иванов",  # ФИО водителя (Отчество Имя Фамилия)
            "+7 900 000-00-00",  # Телефон водителя
            datetime(2026, 1, 15, 9, 0, 0),  # Дата\время подтверждения заявки (МСК)
            "Завершено",  # Статус заявки
            datetime(2026, 1, 15, 10, 0, 0),  # Дата отгрузки (Часовой пояс точки)
            datetime(2026, 1, 15, 18, 0, 0),  # Дата окончания рейса
            15000,  # Сумма транзакций
            0,  # Штраф
        ]
    )

    for i, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(16, len(header) + 2)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Заголовки экспорта (2026-06-29, кнопка "Экспорт" в реестре поездок) - те
# же подписи, что в таблице на экране (Trips.tsx), плюс "Биллинг" между
# "Сумма" и "Штраф" (сумма за вычетом "% СК" перевозчика - см.
# Trips.tsx::billingFor, тот же расчёт, что в commission_pct_for() для
# еженедельного P&L).
EXPORT_HEADERS = [
    "Источник", "Перевозчик", "№ заявки", "Статус", "Отгрузка", "Окончание",
    "Тип", "Водитель", "Машина", "Телефон", "Сумма", "Биллинг", "Штраф",
]


def build_trips_export(rows: list) -> bytes:
    """Реестр поездок -> .xlsx. `rows` (models.TripExportRow) приходит уже
    отфильтрованным и отсортированным с фронтенда, в том же порядке, что
    видно на экране - здесь только сериализация в xlsx, без бизнес-логики
    (фильтры/сортировка/расчёт "Биллинг" не повторяются)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Реестр поездок"
    ws.append(EXPORT_HEADERS)
    for r in rows:
        ws.append(
            [
                r.source, r.carrier_name, r.request_number, r.status, r.dep_at, r.end_at,
                r.tariff_type, r.driver_name, r.truck_label, r.driver_phone,
                r.amount, r.billing, r.fines,
            ]
        )
    for i, header in enumerate(EXPORT_HEADERS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(14, len(header) + 2)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Недельная выгрузка по перевозчику (2026-08-20) ───────────────────────────
# Книга как в присланном пользователем файле «...финал согласно реестрам»:
# «Сводная» + вкладка на каждую неделю отчётности. Рейс попадает в неделю своего
# report_week, штраф — в неделю fines_report_week (может отличаться: у перевозчика
# учёт по неделе ПОСТУПЛЕНИЯ реестра/штрафа, а не по фактической дате рейса).
WEEKLY_EXPORT_HEADERS = [
    "№ заявки", "Внешний № заявки", "Тип тарификации", "гос. Номер ТС",
    "ФИО водителя", "Телефон водителя", "Дата\\время подтверждения заявки (МСК)",
    "Статус заявки", "Дата отгрузки (Часовой пояс точки)", "Дата окончания рейса",
    "Сумма транзакций", "Штраф", "Примечание",
]


def _week_label(monday: date) -> str:
    iso = monday.isocalendar()[1]
    end = monday + timedelta(days=6)
    return f"Н{iso} ({monday.strftime('%d.%m')}-{end.strftime('%d.%m')})"


# Оформление под шаблон пользователя «...финал согласно реестрам»
_HDR_FILL = PatternFill("solid", fgColor="1F3864")      # тёмно-синий заголовок
_HDR_FONT = Font(bold=True, color="FFFFFF")
_TOT_FILL = PatternFill("solid", fgColor="D9E1F2")      # светло-синий «ИТОГ НЕДЕЛИ»
_PARAM_FILL = PatternFill("solid", fgColor="FFF2CC")    # жёлтая ячейка «Ставка»
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_MONEY = "#,##0.00"
_CNT = "0"
_PCT = "0%"
_DTF = "dd\\.mm\\.yyyy\\ hh:mm"


def build_carrier_weekly_export(carrier_name: str, rows: list, income_by_week: Optional[dict] = None, sk_pct: float = 0.0) -> bytes:
    """Книга как в шаблоне пользователя, НА ФОРМУЛАХ (не статика):
    - вкладка на каждую неделю отчётности: рейсы (A–M) + блок «ИТОГ НЕДЕЛИ» в
      столбцах O/P с формулами COUNTIF/SUM по данным этой вкладки;
    - «Сводная» ссылается на итоговые ячейки вкладок (=Нxx!$P$*), считает
      Итого/К выплате через жёлтый параметр «Ставка к выплате» (C3 = 1 − СК%),
      Поступления (по дате платежа) и Накопит. остаток формулами.
    Правка данных в неделе пересчитывает всю сводную.
    rows — dict'ы (см. ключи ниже). income_by_week — {понедельник: сумма}."""
    income_by_week = income_by_week or {}
    weeks: dict = {}

    def wk(monday):
        return weeks.setdefault(monday, {"trips": [], "fines": []})

    for r in rows:
        rw = r.get("report_week")
        fw = r.get("fines_report_week") or rw
        if rw is None:
            continue
        fine = r.get("fines") or 0
        fine_elsewhere = fine and fw is not None and fw != rw
        wk(rw)["trips"].append({**r, "_fine_here": (not fine_elsewhere)})
        if fine_elsewhere:
            wk(fw)["fines"].append(r)

    rate = round(1 - sk_pct / 100, 4)                    # ставка к выплате = доля после СК
    ordered = sorted(set(weeks) | set(income_by_week))
    title_of = {m: _week_label(m)[:31] for m in ordered}

    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "Сводная"

    # ── Вкладки по неделям ──
    for monday in ordered:
        b = weeks.get(monday, {"trips": [], "fines": []})
        ws = wb.create_sheet(title=title_of[monday])
        for ci, h in enumerate(WEEKLY_EXPORT_HEADERS, start=1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = _HDR_FILL; c.font = _HDR_FONT; c.alignment = _CENTER

        rownum = 2

        def _put(vals):
            nonlocal rownum
            for ci, v in enumerate(vals, start=1):
                c = ws.cell(row=rownum, column=ci, value=v)
                if ci in (11, 12):
                    c.number_format = _MONEY
                elif ci in (7, 9, 10) and isinstance(v, datetime):
                    c.number_format = _DTF
            rownum += 1

        for t in b["trips"]:
            _put([
                t.get("request_number"), t.get("external_request_number", ""), t.get("tariff_type", ""),
                t.get("plate", ""), t.get("driver", ""), t.get("driver_phone", ""), t.get("confirmed_at"),
                t.get("status", ""), t.get("dep_at"), t.get("end_at"), t.get("amount") or 0,
                (t.get("fines") or 0) if t["_fine_here"] else None, "",
            ])
        for f in b["fines"]:
            frw = f.get("report_week")
            note = f"Штраф за рейс (неделя рейса {_week_label(frw)})" if frw else "Штраф за рейс"
            # строка-штраф: статус/даты/сумма пустые, чтобы не задваивать рейс в COUNTIF/сумме
            _put([
                f.get("request_number"), f.get("external_request_number", ""), f.get("tariff_type", ""),
                f.get("plate", ""), f.get("driver", ""), f.get("driver_phone", ""), None,
                "", None, None, None, f.get("fines") or 0, note,
            ])

        last = rownum - 1 if rownum > 2 else 1
        ws.freeze_panes = "A2"
        for ci, h in enumerate(WEEKLY_EXPORT_HEADERS, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = max(14, len(h) + 2)

        # Блок «ИТОГ НЕДЕЛИ» (O/P) — на формулах
        oc = ws.cell(row=1, column=15, value="ИТОГ НЕДЕЛИ")
        oc.fill = _TOT_FILL; oc.font = Font(bold=True)
        block = [
            ("Период (неделя)", _week_label(monday), None),
            ("Рейсов выполнено", f'=COUNTIF($H$2:$H${last},"Завершено")', _CNT),
            ("Рейсов отменено", f'=COUNTIF($H$2:$H${last},"Отменено")+COUNTIF($H$2:$H${last},"Срыв")', _CNT),
            ('Прочие статусы («Получен ответ»)', f'=COUNTIF($H$2:$H${last},"Получен ответ")', _CNT),
            ("Всего строк", f'=COUNTA($A$2:$A${last})', _CNT),
            ("Сумма, ₽", f'=SUM($K$2:$K${last})', _MONEY),
            ("Сумма штрафов, ₽", f'=SUM($L$2:$L${last})', _MONEY),
            ("Итого (сумма − штрафы), ₽", "=P7-P8", _MONEY),
            ("К выплате (итого × ставка), ₽", "=P9*Сводная!$C$3", _MONEY),
            ("Поступления (по дате платежа), ₽", round(income_by_week.get(monday, 0) or 0, 2), _MONEY),
        ]
        for i, (label, val, fmt) in enumerate(block, start=2):
            lc = ws.cell(row=i, column=15, value=label)
            if i == 2:
                lc.font = Font(bold=True)
            pc = ws.cell(row=i, column=16, value=val)
            if fmt:
                pc.number_format = fmt
        ws.column_dimensions["O"].width = 34
        ws.column_dimensions["P"].width = 30

    # ── «Сводная» ──
    summary["A1"] = f"Реестр поездок по неделям — перевозчик «{carrier_name}»"
    summary["A1"].font = Font(bold=True, size=13)
    summary["A2"] = "Неделя отчётности = неделя поступления реестра/штрафа к учёту (не факт. дата рейса). Все числа — формулы."
    summary["A3"] = "Ставка к выплате"
    pcell = summary["C3"]; pcell.value = rate; pcell.number_format = _PCT; pcell.fill = _PARAM_FILL; pcell.font = Font(bold=True)
    summary["D3"] = "← жёлтая ячейка: параметр (доля после СК). Пересчитывает «К выплате» на всех вкладках."

    head = ["Период (неделя)", "Вкладка", "Рейсов выполнено", "Рейсов отменено",
            "Прочие статусы", "Всего строк", "Сумма, ₽", "Штрафы, ₽", "Итого, ₽",
            "К выплате (Netto), ₽", "Поступления, ₽", "Накопит. остаток, ₽"]
    HROW = 5
    for ci, h in enumerate(head, start=1):
        c = summary.cell(row=HROW, column=ci, value=h)
        c.fill = _HDR_FILL; c.font = _HDR_FONT; c.alignment = _CENTER

    first = HROW + 1
    r = first
    for monday in ordered:
        t = f"'{title_of[monday]}'"
        summary.cell(row=r, column=1, value=_week_label(monday))
        summary.cell(row=r, column=2, value=title_of[monday])
        for col, pref in ((3, "$P$3"), (4, "$P$4"), (5, "$P$5"), (6, "$P$6"), (7, "$P$7"), (8, "$P$8")):
            cc = summary.cell(row=r, column=col, value=f"={t}!{pref}")
            cc.number_format = _CNT if col <= 6 else _MONEY
        summary.cell(row=r, column=9, value=f"=G{r}-H{r}").number_format = _MONEY
        summary.cell(row=r, column=10, value=f"=I{r}*$C$3").number_format = _MONEY
        summary.cell(row=r, column=11, value=f"={t}!$P$11").number_format = _MONEY
        cum_prev = "" if r == first else f"L{r-1}+"
        summary.cell(row=r, column=12, value=f"={cum_prev}J{r}-K{r}").number_format = _MONEY
        r += 1
    lastr = r - 1

    # ИТОГО
    summary.cell(row=r, column=1, value="ИТОГО").font = Font(bold=True)
    for col in range(3, 12):
        L = get_column_letter(col)
        cc = summary.cell(row=r, column=col, value=f"=SUM({L}{first}:{L}{lastr})")
        cc.number_format = _CNT if col <= 6 else _MONEY
        cc.font = Font(bold=True)
    tc = summary.cell(row=r, column=12, value=f"=L{lastr}")   # накопит остаток = последняя неделя
    tc.number_format = _MONEY; tc.font = Font(bold=True)

    widths = [24, 20, 15, 15, 14, 12, 15, 13, 15, 16, 15, 17]
    for ci, w in enumerate(widths, start=1):
        summary.column_dimensions[get_column_letter(ci)].width = w
    summary.freeze_panes = "A6"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_dep_date(raw) -> date:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    return datetime.strptime(str(raw).strip(), "%d.%m.%Y %H:%M:%S").date()


def _iso_week_monday(d: date) -> date:
    return d - timedelta(days=d.weekday())


def import_trips(file_bytes: bytes, session: Session, source: str = "", carrier_name: str = "", report_week: Optional[date] = None) -> dict:
    """LIVE import path (redesigned 2026-06-19, see module docstring).

    `report_week` (понедельник недели отчётности, из диалога импорта): если
    задан — на КАЖДЫЙ рейс этой партии проставляется report_week, а на рейсы с
    ненулевым штрафом ещё и fines_report_week. Так «неделя загрузки реестра»
    фиксируется для выгрузки по перевозчику. Если None — поля недели не трогаем
    (сохраняются прежние значения, напр. из бэкфилла).

    One Trip row per source row - every status included, no week/driver
    grouping. Re-importing the same (or an updated, overlapping) file is
    safe: rows are upserted by "№ заявки" (request_number), which the
    fixture confirmed is unique per row, since the registry can be
    re-exported as trip statuses change (e.g. "Получен ответ" ->
    "Завершено").

    `source` / `carrier_name` come from the import dialog (chosen by the
    user, not read from the file itself - the file has no such columns)
    and are stamped onto every row touched by this one import call.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Case-insensitive: matching is otherwise exact-string (see module
    # docstring), and a header that differs only by case (e.g. user typed
    # "штраф" instead of "Штраф" when hand-adding the column - found
    # 2026-06-26 testing request 6408388) would silently match nothing and
    # default every row to 0 with no error.
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_index = {str(name).strip().lower(): idx for idx, name in enumerate(header_row) if name}

    def col(row, key):
        idx = header_index.get(TRIP_COLUMNS[key].strip().lower())
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    existing = {t.request_number: t for t in session.exec(select(models.Trip)).all()}
    driver_cache: dict = {}
    truck_cache: dict = {}
    new_driver_names = []
    new_truck_labels = []

    trips_created = 0
    trips_updated = 0
    skipped_bad_rows = 0

    for row in rows:
        if row is None or all(c is None for c in row):
            continue

        request_number = _cell_str(col(row, "request_number"))
        if not request_number:
            skipped_bad_rows += 1
            continue

        try:
            dep_at = _parse_dt(col(row, "dep_at"))
        except (ValueError, TypeError):
            dep_at = None
        if dep_at is None:
            skipped_bad_rows += 1
            continue

        try:
            confirmed_at = _parse_dt(col(row, "confirmed_at"))
        except (ValueError, TypeError):
            confirmed_at = None
        try:
            end_at = _parse_dt(col(row, "end_at"))
        except (ValueError, TypeError):
            end_at = None

        driver_fio = _cell_str(col(row, "driver"))
        plate_raw = _cell_str(col(row, "plate"))

        driver_id = None
        if driver_fio:
            surname = _surname_token(driver_fio)
            if surname not in driver_cache:
                driver, was_created = find_or_create_driver(session, driver_fio)
                driver_cache[surname] = driver
                if was_created:
                    new_driver_names.append(driver.name)
            driver_id = driver_cache[surname].id

        truck_id = None
        if plate_raw:
            norm_plate = normalize_plate(plate_raw)
            if norm_plate not in truck_cache:
                truck, was_created = find_or_create_truck(session, plate_raw)
                truck_cache[norm_plate] = truck
                if was_created:
                    new_truck_labels.append(truck.label)
            truck_id = truck_cache[norm_plate].id

        fields = dict(
            request_number=request_number,
            external_request_number=_cell_str(col(row, "external_request_number")),
            tariff_type=_cell_str(col(row, "tariff_type")),
            plate_raw=plate_raw,
            truck_id=truck_id,
            driver_name_raw=driver_fio,
            driver_id=driver_id,
            driver_phone=_cell_str(col(row, "driver_phone")),
            confirmed_at=confirmed_at,
            status=_cell_str(col(row, "status")),
            dep_at=dep_at,
            end_at=end_at,
            amount=col(row, "amount") or 0,
            fines=col(row, "fines") or 0,
            source=source,
            carrier_name=carrier_name,
        )
        # Неделя отчётности из диалога импорта: рейсу — report_week, рейсу со
        # штрафом — ещё и fines_report_week (штраф поступил к учёту этой неделей).
        if report_week is not None:
            fields["report_week"] = report_week
            if (fields["fines"] or 0) > 0:
                fields["fines_report_week"] = report_week

        trip = existing.get(request_number)
        if trip:
            for k, v in fields.items():
                # Не затираем ненулевую выручку/штраф нулём из следующей выгрузки:
                # штрафы приходят позже отдельным файлом где amount=0, и наоборот.
                if k in ("amount", "fines") and (v or 0) == 0 and (getattr(trip, k) or 0) > 0:
                    continue
                setattr(trip, k, v)
            trips_updated += 1
        else:
            trip = models.Trip(**fields)
            existing[request_number] = trip
            trips_created += 1
        session.add(trip)

    session.commit()

    return {
        "total_rows": len(rows),
        "trips_created": trips_created,
        "trips_updated": trips_updated,
        "skipped_bad_rows": skipped_bad_rows,
        "new_drivers": new_driver_names,
        "new_trucks": new_truck_labels,
    }


def import_trip_registry(file_bytes: bytes, session: Session) -> dict:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_index = {str(name).strip(): idx for idx, name in enumerate(header_row) if name}

    def col(row, key):
        idx = header_index.get(EXPECTED_COLUMNS[key])
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    skipped_cancelled = 0
    skipped_bad_rows = 0
    parsed = []

    for row in rows:
        if row is None or all(c is None for c in row):
            continue
        status = col(row, "status")
        if status in STATUS_EXCLUDE:
            skipped_cancelled += 1
            continue
        if not status:
            skipped_bad_rows += 1
            continue
        try:
            dep_date = _parse_dep_date(col(row, "dep_date"))
        except (ValueError, TypeError):
            skipped_bad_rows += 1
            continue
        driver_fio = (col(row, "driver") or "").strip()
        if not driver_fio:
            skipped_bad_rows += 1
            continue
        parsed.append(
            {
                "week": _iso_week_monday(dep_date),
                "surname": _surname_token(driver_fio),
                "fio": driver_fio,
                "plate": (col(row, "plate") or "").strip(),
                "amount": col(row, "amount") or 0,
                "date": dep_date,
                "trip_type": (col(row, "trip_type") or "").strip(),
            }
        )

    groups = defaultdict(list)
    for item in parsed:
        groups[(item["week"], item["surname"])].append(item)

    driver_cache: dict = {}
    truck_cache: dict = {}
    new_driver_names = []
    new_truck_labels = []
    batches_created = 0

    for (_week_key, surname), items in groups.items():
        if surname not in driver_cache:
            driver, created = find_or_create_driver(session, items[0]["fio"])
            driver_cache[surname] = driver
            if created:
                new_driver_names.append(driver.name)
        driver = driver_cache[surname]

        plate_counts = Counter(i["plate"] for i in items if i["plate"])
        main_plate_raw = plate_counts.most_common(1)[0][0] if plate_counts else ""
        norm_plate = normalize_plate(main_plate_raw)
        if norm_plate not in truck_cache:
            truck, created = find_or_create_truck(session, main_plate_raw)
            truck_cache[norm_plate] = truck
            if created:
                new_truck_labels.append(truck.label)
        truck = truck_cache[norm_plate]

        dates = [i["date"] for i in items]
        gross = sum(i["amount"] for i in items)
        trips = len(items)
        trip_types = {i["trip_type"] for i in items if i["trip_type"]}
        route_name = trip_types.pop() if len(trip_types) == 1 else ", ".join(sorted(trip_types))

        batch = models.TripBatch(
            period_start=min(dates),
            period_end=max(dates),
            driver_id=driver.id,
            truck_id=truck.id,
            route_id=None,
            km_per_trip=0,
            rate_per_trip=round(gross / trips, 2) if trips else 0,
            trips_count=trips,
            gross_revenue=gross,
            fines=0,
            toll_road=0,
            fuel_cost=0,
            commission_pct=None,
            route_name=route_name,
        )
        session.add(batch)
        batches_created += 1

    session.commit()

    return {
        "total_rows": len(rows),
        "batches_created": batches_created,
        "skipped_cancelled": skipped_cancelled,
        "skipped_bad_rows": skipped_bad_rows,
        "new_drivers": new_driver_names,
        "new_trucks": new_truck_labels,
    }

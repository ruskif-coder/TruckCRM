"""Баланс перевозчиков (2026-07-12, v1.1.3).

Детализация по неделям:
  gross       = Σ trip.amount   (не отменённые рейсы перевозчика за неделю)
  fines       = Σ trip.fines    (штрафы из того же отчёта, колонка «Штраф»)
  net         = (gross - fines) × (1 - carrier.insurance_pct / 100)

Накопительный баланс:
  paid        = Σ CashFlowEntry.income  где counterparty совпадает с именем
                контрагента, привязанного к перевозчику (Carrier.counterparty_id)
  balance     = Σ net_week - paid
"""

from collections import defaultdict
from datetime import date as date_type, datetime, timedelta
from io import BytesIO

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from sqlmodel import Session, select

from .. import models
from ..auth import get_current_user
from ..database import get_session
from ..importers.trip_registry import EXPORT_HEADERS, build_carrier_weekly_export
from ..models import Carrier, CashFlowEntry, Counterparty, Driver, Trip, Truck

router = APIRouter(prefix="/api/carriers/balance", tags=["carrier-balance"])
_auth = [Depends(get_current_user)]

_STAFF_ROLES = {"admin", "foreman", "accountant"}


def _require_staff(user: models.User = Depends(get_current_user)) -> models.User:
    """Финансовые данные перевозчиков — только для staff (аудит-2026-07-13)."""
    if user.role not in _STAFF_ROLES:
        raise HTTPException(status_code=403, detail="Доступ только для admin/foreman/accountant")
    return user


def _iso_week_monday(d: date_type) -> date_type:
    return d - timedelta(days=d.weekday())


def _as_date(v) -> date_type:
    """CashFlowEntry.date/Trip.dep_at могут прийти как date или datetime — нормализуем."""
    return v.date() if isinstance(v, datetime) else v


def _round2(v: float) -> float:
    return round(v, 2)


@router.get("/")
def carrier_balance_summary(
    session: Session = Depends(get_session),
    _user: models.User = Depends(_require_staff),
):
    """Сводка баланса по всем перевозчикам (за всё время, накопительно)."""
    carriers = session.exec(select(Carrier)).all()
    counterparties = {c.id: c for c in session.exec(select(Counterparty)).all()}
    trips = session.exec(select(Trip)).all()
    cashflow = session.exec(select(CashFlowEntry)).all()

    # Индекс: carrier_name -> carrier
    carrier_by_name: dict[str, Carrier] = {(c.name or "").strip(): c for c in carriers}

    # Обратный индекс: counterparty_name -> list[carrier_name] (для матчинга платежей)
    cp_name_to_carriers: dict[str, list[str]] = defaultdict(list)
    for c in carriers:
        if c.counterparty_id and c.counterparty_id in counterparties:
            cp_name = (counterparties[c.counterparty_id].name or "").strip()
            if cp_name:
                cp_name_to_carriers[cp_name].append((c.name or "").strip())

    # Рейсы: группируем по (carrier_name, ISO-неделя dep_at)
    # gross и trips  — только не отменённые (выручки нет)
    # fines          — ВСЕ рейсы, включая отменённые: штраф за отмену
    #                  реален и должен уменьшать то, что мы платим перевозчику
    week_buckets: dict = defaultdict(lambda: {"gross": 0.0, "fines": 0.0, "trips": 0})
    for t in trips:
        if not t.dep_at:
            continue
        name = (t.carrier_name or t.source or "").strip()
        if not name:
            continue
        wk = _iso_week_monday(t.dep_at.date())
        key = (name, wk)
        cancelled = (t.status or "").lower().startswith("отмен")
        if not cancelled:
            week_buckets[key]["gross"] += t.amount or 0
            week_buckets[key]["trips"] += 1
        week_buckets[key]["fines"] += t.fines or 0  # штрафы — всегда

    # Поступления по неделям: CashFlowEntry.income от контрагента перевозчика,
    # ПРИВЯЗКА К ДАТЕ ПЛАТЕЖА (entry.date), а не к неделе рейса.
    carrier_paid: dict[str, float] = defaultdict(float)            # всего (накопительно)
    carrier_week_income: dict = defaultdict(float)                 # (name, wk) -> сумма
    for entry in cashflow:
        if not (entry.income and entry.income > 0):
            continue
        cp_text = (entry.counterparty or "").strip()
        names = cp_name_to_carriers.get(cp_text, [])
        wk = _iso_week_monday(_as_date(entry.date)) if entry.date else None
        for carrier_name in names:
            carrier_paid[carrier_name] += entry.income
            if wk is not None:
                carrier_week_income[(carrier_name, wk)] += entry.income

    # Агрегируем по неделям — объединяем недели рейсов и недели поступлений.
    carrier_net: dict[str, float] = defaultdict(float)
    carrier_gross: dict[str, float] = defaultdict(float)
    carrier_fines: dict[str, float] = defaultdict(float)
    carrier_trips: dict[str, int] = defaultdict(int)
    carrier_week_map: dict[str, dict] = defaultdict(dict)          # name -> {wk: week_dict}

    def _week_slot(name, wk):
        slot = carrier_week_map[name].get(wk)
        if slot is None:
            slot = {
                "week_start": wk.isoformat(), "week_end": (wk + timedelta(days=6)).isoformat(),
                "trips": 0, "gross": 0.0, "fines": 0.0, "net": 0.0, "income": 0.0,
            }
            carrier_week_map[name][wk] = slot
        return slot

    for (name, wk), g in week_buckets.items():
        carrier = carrier_by_name.get(name)
        sk_pct = (carrier.insurance_pct or 0.0) if carrier else 0.0
        gross = g["gross"]; fines = g["fines"]
        net = (gross - fines) * (1 - sk_pct / 100)
        carrier_net[name] += net
        carrier_gross[name] += gross
        carrier_fines[name] += fines
        carrier_trips[name] += g["trips"]
        slot = _week_slot(name, wk)
        slot.update(trips=g["trips"], gross=_round2(gross), fines=_round2(fines), net=_round2(net))

    for (name, wk), inc in carrier_week_income.items():
        _week_slot(name, wk)["income"] = _round2(inc)

    # Собираем итог — перевозчики с рейсами + перевозчики с поступлениями
    processed_names: set = set(carrier_gross.keys()) | set(carrier_paid.keys())

    result = []
    for name in sorted(processed_names):
        carrier = carrier_by_name.get(name)
        cp_id = carrier.counterparty_id if carrier else None
        cp = counterparties.get(cp_id) if cp_id else None
        gross = carrier_gross.get(name, 0.0)
        fines = carrier_fines.get(name, 0.0)
        net = carrier_net.get(name, 0.0)
        paid = carrier_paid.get(name, 0.0)
        balance = net - paid
        weeks = sorted(carrier_week_map.get(name, {}).values(), key=lambda w: w["week_start"])
        result.append({
            "carrier_name": name,
            "carrier_id": carrier.id if carrier else None,
            "counterparty_id": cp_id,
            "counterparty_name": cp.name if cp else None,
            "trips": carrier_trips.get(name, 0),
            "gross": _round2(gross),
            "fines": _round2(fines),
            "net": _round2(net),
            "paid": _round2(paid),
            "balance": _round2(balance),
            "weeks": weeks,
        })

    result.sort(key=lambda r: -r["balance"])
    return result


@router.get("/export")
def carrier_export(
    carrier: str = Query(..., description="Имя перевозчика"),
    session: Session = Depends(get_session),
    user: models.User = Depends(_require_staff),
):
    """XLSX по одному перевозчику (кнопка выгрузки в строке «Отчёты →
    Перевозчики»), 2 листа:
      1) «Сводная по неделям» — итоги за всё время (виджеты) + разбивка по неделям
         (как в раскрывающемся отчёте: Рейсов/Брутто/Штрафы/Netto после СК);
      2) «Реестр рейсов» — все рейсы перевозчика за всё время в нашем стандартном
         формате (те же колонки, что и экспорт реестра поездок).
    """
    name = carrier.strip()
    row = next((r for r in carrier_balance_summary(session, user) if r["carrier_name"] == name), None)
    if row is None:
        raise HTTPException(404, "Перевозчик не найден")

    wb = openpyxl.Workbook()

    # ── Лист 1: сводная по неделям + итоги за всё время ──
    ws1 = wb.active
    ws1.title = "Сводная по неделям"
    ws1.append([f"Перевозчик: {name}"])
    ws1.append([])
    ws1.append(["Итоги за всё время", ""])
    ws1.append(["Рейсов", row["trips"]])
    ws1.append(["Брутто", row["gross"]])
    ws1.append(["Штрафы", row["fines"]])
    ws1.append(["Netto (после СК)", row["net"]])
    ws1.append(["Оплачено", row["paid"]])
    ws1.append(["Баланс", row["balance"]])
    ws1.append([])
    ws1.append(["Неделя", "Рейсов", "Брутто", "Штрафы", "Netto (после СК)"])
    for w in row["weeks"]:
        ws1.append([f'{w["week_start"]} – {w["week_end"]}', w["trips"], w["gross"], w["fines"], w["net"]])
    ws1.append(["ИТОГО", row["trips"], row["gross"], row["fines"], row["net"]])
    for i, wd in enumerate([26, 12, 14, 12, 18], start=1):
        ws1.column_dimensions[ws1.cell(row=1, column=i).column_letter].width = wd

    # ── Лист 2: реестр рейсов за всё время (стандартный формат) ──
    ws2 = wb.create_sheet("Реестр рейсов")
    ws2.append(EXPORT_HEADERS)
    drivers = {d.id: d for d in session.exec(select(Driver)).all()}
    trucks = {t.id: t for t in session.exec(select(Truck)).all()}
    carrier_obj = next((c for c in session.exec(select(Carrier)).all() if (c.name or "").strip() == name), None)
    sk = (carrier_obj.insurance_pct or 0.0) if carrier_obj else 0.0
    ctrips = [t for t in session.exec(select(Trip)).all() if (t.carrier_name or t.source or "").strip() == name]
    ctrips.sort(key=lambda t: t.dep_at or datetime.min)

    def _fmt(dt):
        return dt.strftime("%d.%m.%Y %H:%M:%S") if dt else ""

    for t in ctrips:
        drv = drivers.get(t.driver_id)
        trk = trucks.get(t.truck_id)
        driver_name = (drv.name if drv else "") or t.driver_name_raw or ""
        truck_label = (trk.plate if trk else "") or t.plate_raw or ""
        billing = (t.amount or 0) * (1 - sk / 100)   # Биллинг = сумма минус % СК
        ws2.append([
            t.source or "", t.carrier_name or "", t.request_number or "", t.status or "",
            _fmt(t.dep_at), _fmt(t.end_at), t.tariff_type or "", driver_name, truck_label,
            t.driver_phone or "", t.amount or 0, _round2(billing), t.fines or 0,
        ])
    for i, header in enumerate(EXPORT_HEADERS, start=1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = max(14, len(header) + 2)

    buf = BytesIO()
    wb.save(buf)
    safe = "".join(ch for ch in name if ch.isascii() and (ch.isalnum() or ch in " _-")).strip()[:40] or "perevozchik"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="carrier_{safe}.xlsx"'},
    )


@router.get("/weekly-export")
def carrier_weekly_export(
    carrier: str = Query(..., description="Имя перевозчика"),
    session: Session = Depends(get_session),
    _user: models.User = Depends(_require_staff),
):
    """XLSX по перевозчику «как в реестре перевозчика»: «Сводная» + вкладка на
    каждую неделю отчётности. Рейс попадает в неделю report_week, штраф — в
    неделю fines_report_week (учёт перевозчика по неделе поступления реестра, а
    не по факт. дате). Недели проставляются импортом («Неделя отчётности») или
    историческим бэкфиллом (/api/maintenance/backfill-report-weeks)."""
    name = carrier.strip()
    drivers = {d.id: d for d in session.exec(select(Driver)).all()}
    trucks = {t.id: t for t in session.exec(select(Truck)).all()}
    ctrips = [t for t in session.exec(select(Trip)).all() if (t.carrier_name or t.source or "").strip() == name]
    if not ctrips:
        raise HTTPException(404, "Перевозчик не найден или нет рейсов")

    rows = []
    for t in ctrips:
        drv = drivers.get(t.driver_id)
        trk = trucks.get(t.truck_id)
        # Неделя отчётности: если не проставлена (нет бэкфилла/импорта с неделей) —
        # фолбэк на неделю по дате отгрузки/окончания, чтобы книга не была пустой.
        base = t.dep_at or t.end_at
        fallback = _iso_week_monday(base.date()) if base else None
        rw = t.report_week or fallback
        fw = t.fines_report_week or rw
        rows.append({
            "request_number": t.request_number, "external_request_number": t.external_request_number or "",
            "tariff_type": t.tariff_type or "", "plate": (trk.plate if trk else "") or t.plate_raw or "",
            "driver": (drv.name if drv else "") or t.driver_name_raw or "", "driver_phone": t.driver_phone or "",
            "confirmed_at": t.confirmed_at, "status": t.status or "", "dep_at": t.dep_at, "end_at": t.end_at,
            "amount": t.amount or 0, "fines": t.fines or 0,
            "report_week": rw, "fines_report_week": fw,
        })
    # Поступления по неделям от привязанного контрагента (по дате платежа).
    carrier_obj = next((c for c in session.exec(select(Carrier)).all() if (c.name or "").strip() == name), None)
    cp_name = None
    if carrier_obj and carrier_obj.counterparty_id:
        cp = session.get(Counterparty, carrier_obj.counterparty_id)
        cp_name = (cp.name or "").strip() if cp else None
    income_by_week: dict = {}
    if cp_name:
        for e in session.exec(select(CashFlowEntry)).all():
            if e.income and e.income > 0 and (e.counterparty or "").strip() == cp_name and e.date:
                wk = _iso_week_monday(_as_date(e.date))
                income_by_week[wk] = income_by_week.get(wk, 0) + e.income

    sk_pct = (carrier_obj.insurance_pct or 0.0) if carrier_obj else 0.0
    data = build_carrier_weekly_export(name, rows, income_by_week, sk_pct)
    safe = "".join(ch for ch in name if ch.isascii() and (ch.isalnum() or ch in " _-")).strip()[:40] or "perevozchik"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="carrier_weekly_{safe}.xlsx"'},
    )

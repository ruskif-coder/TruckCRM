"""Техническое обслуживание системы (admin-only, 2026-07-05).

Эндпойнты для ручного или планового обслуживания:
  POST /api/maintenance/cleanup-photos  — удалить осиротевшие фото из PHOTOS_DIR

Принципы:
- Только admin (require_role("admin")), эндпойнт не в конфигурируемой матрице ролей.
- Никогда не удаляет файлы, на которые есть ссылки в БД — только «осиротевшие».
- Сообщает о файлах старше 12 месяцев (даже если они привязаны к актам) —
  решение об удалении остаётся за администратором.
"""
import json
import os
import re
from datetime import date, datetime, timedelta
from io import BytesIO

import openpyxl
from fastapi import APIRouter, Depends, File, UploadFile
from sqlmodel import Session, select

from .. import models
from ..auth import require_role
from ..database import get_session
from ..importers.trip_registry import TRIP_COLUMNS, _cell_str

router = APIRouter(prefix="/api/maintenance", tags=["maintenance"])
_admin = [Depends(require_role("admin"))]


def _iso_monday(d: date) -> date:
    return d - timedelta(days=d.weekday())

# Порог «старых» файлов — только информируем, не удаляем
_PHOTO_MAX_AGE_DAYS = 365


@router.post("/cleanup-photos", dependencies=_admin)
def cleanup_photos(session: Session = Depends(get_session)):
    """Очищает папку фото (PHOTOS_DIR) от осиротевших файлов.

    «Осиротевший» — файл, которого нет ни в одной из трёх таблиц:
      • InspectionDamage.photo_path     (приёмка авто)
      • RepairRequest.photo_paths       (JSON-список, заявки на ремонт)
      • CompensationRequest.photo_paths (JSON-список, заявки на компенсацию)

    Файлы, привязанные к записям БД, не удаляются независимо от возраста —
    только попадают в список old_referenced_files для информирования.

    Возвращает:
      orphaned_deleted       — удалено осиротевших файлов
      orphaned_errors        — ошибки при удалении (нет прав и т.п.)
      old_referenced_count   — файлов, привязанных к БД, но старше 12 мес.
      old_referenced_files   — [{filename, mtime}] для ревью
      total_files_before     — сколько файлов было в папке до очистки
    """
    photos_dir = os.environ.get("PHOTOS_DIR", "./photos")

    if not os.path.isdir(photos_dir):
        return {
            "status": "photos_dir_not_found",
            "dir": photos_dir,
            "orphaned_deleted": 0,
            "orphaned_errors": 0,
            "old_referenced_count": 0,
            "old_referenced_files": [],
            "total_files_before": 0,
        }

    # 1. Собираем все имена файлов, на которые есть ссылки в БД
    referenced: set[str] = set()

    for dmg in session.exec(select(models.InspectionDamage)).all():
        if dmg.photo_path:
            referenced.add(dmg.photo_path)

    for req in session.exec(select(models.RepairRequest)).all():
        if req.photo_paths:
            try:
                for fn in json.loads(req.photo_paths):
                    if fn:
                        referenced.add(fn)
            except (json.JSONDecodeError, TypeError):
                pass

    for comp in session.exec(select(models.CompensationRequest)).all():
        if comp.photo_paths:
            try:
                for fn in json.loads(comp.photo_paths):
                    if fn:
                        referenced.add(fn)
            except (json.JSONDecodeError, TypeError):
                pass

    # 2. Сканируем папку и разбираем по категориям
    cutoff = datetime.now() - timedelta(days=_PHOTO_MAX_AGE_DAYS)
    orphaned_deleted = 0
    orphaned_errors = 0
    old_referenced: list[dict] = []
    total_files_before = 0

    for filename in os.listdir(photos_dir):
        filepath = os.path.join(photos_dir, filename)
        if not os.path.isfile(filepath):
            continue
        total_files_before += 1

        try:
            mtime = datetime.fromtimestamp(os.path.getmtime(filepath))
        except OSError:
            mtime = datetime.now()

        is_referenced = filename in referenced

        if not is_referenced:
            # Осиротевший файл — удаляем
            try:
                os.remove(filepath)
                orphaned_deleted += 1
            except OSError:
                orphaned_errors += 1
        elif mtime < cutoff:
            # Привязан к БД, но старше порога — информируем
            old_referenced.append({
                "filename": filename,
                "mtime": mtime.strftime("%Y-%m-%d"),
                "age_days": (datetime.now() - mtime).days,
            })

    note_parts = [f"Удалено {orphaned_deleted} осиротевших файлов."]
    if orphaned_errors:
        note_parts.append(f"Ошибок удаления: {orphaned_errors}.")
    if old_referenced:
        note_parts.append(
            f"Найдено {len(old_referenced)} файлов старше {_PHOTO_MAX_AGE_DAYS} дн., "
            "привязанных к актам/заявкам — не удалены. Просмотрите список и решите вручную."
        )
    else:
        note_parts.append(f"Файлов старше {_PHOTO_MAX_AGE_DAYS} дн. не обнаружено.")

    return {
        "status": "ok",
        "orphaned_deleted": orphaned_deleted,
        "orphaned_errors": orphaned_errors,
        "old_referenced_count": len(old_referenced),
        "old_referenced_files": old_referenced,
        "total_files_before": total_files_before,
        "note": " ".join(note_parts),
    }


@router.post("/backfill-report-weeks", dependencies=_admin)
async def backfill_report_weeks(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
):
    """Исторический бэкфилл недель отчётности из недельного файла-реестра
    («...по неделям... финал согласно реестрам»): для каждого рейса по № заявки
    проставляет report_week (неделя вкладки с его строкой рейса) и
    fines_report_week (неделя вкладки со строкой-штрафом, если штраф вынесен в
    другую неделю). Неделя вкладки берётся из фактических дат её строк (год —
    из данных, не из названия). Идемпотентно, только обновляет существующие рейсы."""
    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)

    report_week: dict[str, date] = {}
    fines_week: dict[str, date] = {}
    sheets = 0
    for name in wb.sheetnames:
        if not (name.strip().startswith("Н") and "(" in name):
            continue
        rows = list(wb[name].iter_rows(values_only=True))
        if not rows:
            continue
        hidx = {str(h).strip().lower(): i for i, h in enumerate(rows[0]) if h}

        def col(r, key, hidx=hidx):
            i = hidx.get(TRIP_COLUMNS[key].strip().lower())
            return r[i] if (i is not None and i < len(r)) else None

        # Понедельник недели вкладки — из первой строки с реальной датой рейса.
        monday = None
        for r in rows[1:]:
            for dv in (col(r, "dep_at"), col(r, "end_at")):
                if isinstance(dv, datetime):
                    monday = _iso_monday(dv.date())
                    break
                if isinstance(dv, date):
                    monday = _iso_monday(dv)
                    break
            if monday:
                break
        if monday is None:
            continue
        sheets += 1

        for r in rows[1:]:
            rn = _cell_str(col(r, "request_number"))
            if not rn:
                continue
            status = _cell_str(col(r, "status"))
            dep = col(r, "dep_at")
            end = col(r, "end_at")
            fine = col(r, "fines")
            is_trip = bool(status) or dep is not None or end is not None
            is_fine = fine not in (None, 0, "")
            if is_trip and rn not in report_week:
                report_week[rn] = monday
            if is_fine and not is_trip:
                fines_week[rn] = monday
            elif is_fine and is_trip:
                fines_week.setdefault(rn, monday)

    # Применяем к существующим рейсам
    updated_report = 0
    updated_fines = 0
    fines_separate = 0
    for t in session.exec(select(models.Trip)).all():
        rn = str(t.request_number)
        if rn in report_week and t.report_week != report_week[rn]:
            t.report_week = report_week[rn]
            updated_report += 1
            session.add(t)
        fw = fines_week.get(rn) or report_week.get(rn)
        if fw and t.fines_report_week != fw:
            t.fines_report_week = fw
            updated_fines += 1
            session.add(t)
        if rn in fines_week and rn in report_week and fines_week[rn] != report_week[rn]:
            fines_separate += 1
    session.commit()

    return {
        "status": "ok",
        "sheets_parsed": sheets,
        "requests_in_file": len(report_week),
        "trips_report_week_set": updated_report,
        "trips_fines_week_set": updated_fines,
        "fines_in_separate_week": fines_separate,
    }
